"""
COLEXA BIOSENSOR - HVAC Monitoring Platform
ReportLab PDF and OpenPyXL Excel export engines.

Generated files are written to the exports/ directory with a timestamped
filename so repeated exports never collide or silently overwrite prior
compliance evidence.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage,
)

EXPORTS_DIR: str = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
ASSETS_DIR: str = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets")

BRAND_CYAN = colors.HexColor("#06B6D4")
BRAND_GREEN = colors.HexColor("#C1F24D")
BRAND_DARK = colors.HexColor("#111827")
BRAND_RED = colors.HexColor("#EF4444")


def _ensure_exports_dir() -> None:
    """Create the exports directory if it does not already exist."""
    os.makedirs(EXPORTS_DIR, exist_ok=True)


def generate_compliance_pdf(facility_df: pd.DataFrame, deviation_df: pd.DataFrame, report_title: str = "HVAC Compliance Report") -> str | None:
    """Generate a branded PDF compliance report.

    Args:
        facility_df: DataFrame of recent facility_logs rows to include.
        deviation_df: DataFrame of recent deviation_logs rows to include.
        report_title: Title text shown at the top of the report.

    Returns:
        Absolute path to the generated PDF file, or None on failure.
    """
    try:
        _ensure_exports_dir()
        timestamp_tag: str = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path: str = os.path.join(EXPORTS_DIR, f"COLEXA_HVAC_Compliance_{timestamp_tag}.pdf")

        doc = SimpleDocTemplate(
            file_path, pagesize=A4,
            leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("ColexaTitle", parent=styles["Title"], textColor=BRAND_DARK, fontSize=18)
        subtitle_style = ParagraphStyle("ColexaSub", parent=styles["Normal"], textColor=colors.HexColor("#334155"), fontSize=9)
        section_style = ParagraphStyle("ColexaSection", parent=styles["Heading2"], textColor=BRAND_DARK, fontSize=13, spaceBefore=10)

        story: list = []

        logo_path: str = os.path.join(ASSETS_DIR, "logo.jpg")
        if os.path.exists(logo_path):
            try:
                story.append(RLImage(logo_path, width=32 * mm, height=16 * mm))
                story.append(Spacer(1, 6))
            except Exception:
                pass

        story.append(Paragraph("COLEXA BIOSENSOR", title_style))
        story.append(Paragraph(report_title, ParagraphStyle("Sub2", parent=styles["Heading3"], textColor=BRAND_CYAN)))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &nbsp;|&nbsp; "
            "Regulatory Context: ISO 13485:2016 / FDA 21 CFR Part 11 baseline &nbsp;|&nbsp; "
            "Reference Log: CBL/ENG/02/R01",
            subtitle_style,
        ))
        story.append(Spacer(1, 10))

        story.append(Paragraph("Shift Telemetry Summary", section_style))
        if facility_df.empty:
            story.append(Paragraph("No facility log entries available for this period.", styles["Normal"]))
        else:
            display_cols = [c for c in [
                "shift_date", "shift_time", "ahu_temperature", "ahu_rh",
                "dhu1_rh", "dhu2_rh", "compressor_pressure", "status",
            ] if c in facility_df.columns]
            table_data = [display_cols] + facility_df[display_cols].astype(str).values.tolist()
            report_table = Table(table_data, repeatRows=1)
            report_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#94A3B8")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
            ]))
            story.append(report_table)

        story.append(Spacer(1, 14))
        story.append(Paragraph("Deviation / CAPA Log", section_style))
        if deviation_df.empty:
            story.append(Paragraph("No deviations recorded for this period.", styles["Normal"]))
        else:
            dev_cols = [c for c in [
                "timestamp", "equipment", "parameter", "observed_value",
                "severity", "referenced_sop", "capa_status",
            ] if c in deviation_df.columns]
            dev_data = [dev_cols] + deviation_df[dev_cols].astype(str).values.tolist()
            dev_table = Table(dev_data, repeatRows=1)
            dev_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_RED),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#94A3B8")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FEF2F2")]),
            ]))
            story.append(dev_table)

        story.append(Spacer(1, 16))
        story.append(Paragraph(
            "This report is system-generated by the COLEXA HVAC Monitoring Platform and forms part of the "
            "facility's electronic record-keeping under 21 CFR Part 11 baseline controls. Manual review and "
            "sign-off by Engineering/QA is required before use as a formal batch record attachment.",
            ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#64748B")),
        ))

        doc.build(story)
        return file_path
    except Exception:
        return None


def generate_excel_export(facility_df: pd.DataFrame, deviation_df: pd.DataFrame, file_prefix: str = "COLEXA_HVAC_Export") -> str | None:
    """Generate a branded multi-sheet Excel workbook export.

    Args:
        facility_df: DataFrame of facility_logs rows to include.
        deviation_df: DataFrame of deviation_logs rows to include.
        file_prefix: Filename prefix (timestamp is appended automatically).

    Returns:
        Absolute path to the generated .xlsx file, or None on failure.
    """
    try:
        _ensure_exports_dir()
        timestamp_tag: str = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path: str = os.path.join(EXPORTS_DIR, f"{file_prefix}_{timestamp_tag}.xlsx")

        workbook = Workbook()
        header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        def _write_sheet(ws, dataframe: pd.DataFrame, title: str) -> None:
            ws.title = title
            if dataframe.empty:
                ws.append(["No data available"])
                return
            ws.append(list(dataframe.columns))
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in dataframe.itertuples(index=False):
                ws.append(list(row))
            for idx, column in enumerate(dataframe.columns, start=1):
                max_len = max([len(str(column))] + [len(str(v)) for v in dataframe[column].astype(str).values]) if len(dataframe) else len(str(column))
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 3, 40)

        ws1 = workbook.active
        _write_sheet(ws1, facility_df, "Facility Logs")

        ws2 = workbook.create_sheet("Deviations")
        _write_sheet(ws2, deviation_df, "Deviations")

        workbook.save(file_path)
        return file_path
    except Exception:
        return None
